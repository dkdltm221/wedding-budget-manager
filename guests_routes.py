import io
from pathlib import Path
from typing import Optional
from flask import Blueprint, render_template, request, redirect, url_for, flash, send_file
from models import db, Guest
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font
from fpdf import FPDF
from services.guests_service import (
    sorting_from_request,
    search_filter,
    fetch_guests,
    totals,
    parse_amount,
    create_guest
)
from services.guests_view import build_guest_page_context
from forms import GuestForm

guests_bp = Blueprint('guests', __name__)
HIGHLIGHT_THRESHOLD = 300000  # 금액 배지 표시 기준 (원)


@guests_bp.route('/', methods=['GET', 'POST'])
def list_guests():
    """손님 목록, 추가, 검색, 정렬을 처리한다."""
    form = GuestForm(request.form, meta={'csrf': False})
    if request.method == 'POST':
        if form.validate():
            try:
                amount_value = parse_amount(form.amount.data)
            except ValueError:
                flash('금액을 숫자로 입력해주세요.', 'danger')
            else:
                create_guest(name=form.name.data.strip(), side=form.side.data.strip(),
                             amount_value=amount_value, note=form.note.data or "")
                flash('축의금이 추가되었습니다.', 'success')
                return redirect(url_for('guests.list_guests'))
        else:
            flash('이름, 소속, 금액은 필수이며 금액은 0 이상이어야 합니다.', 'danger')

    context = build_guest_page_context(request.args)
    context['highlight_threshold'] = HIGHLIGHT_THRESHOLD
    context['form'] = form
    return render_template('guests.html', **context)


@guests_bp.route('/<int:guest_id>/edit', methods=['GET', 'POST'])
def edit_guest(guest_id):
    """손님 정보를 수정한다."""
    guest = Guest.query.get_or_404(guest_id)

    if request.method == 'POST':
        name = request.form.get('name', '').strip()
        side = request.form.get('side', '').strip()
        amount = request.form.get('amount', '').strip()
        note = request.form.get('note', '').strip()

        if not name or not side or not amount:
            flash('이름, 소속, 금액은 필수입니다.', 'danger')
            return redirect(url_for('guests.edit_guest', guest_id=guest.id))

        try:
            amount_value = int(amount)
        except ValueError:
            flash('금액을 숫자로 입력해주세요.', 'danger')
            return redirect(url_for('guests.edit_guest', guest_id=guest.id))

        guest.name = name
        guest.side = side
        guest.amount = amount_value
        guest.note = note

        db.session.commit()
        flash('축의금이 수정되었습니다.', 'success')
        return redirect(url_for('guests.list_guests'))

    return render_template('guests_edit.html', guest=guest)


@guests_bp.route('/<int:guest_id>/delete', methods=['POST'])
def delete_guest(guest_id):
    """손님 정보를 삭제한다."""
    guest = Guest.query.get_or_404(guest_id)
    db.session.delete(guest)
    db.session.commit()
    flash('축의금이 삭제되었습니다.', 'info')
    return redirect(url_for('guests.list_guests'))


@guests_bp.route('/export/excel')
def export_excel():
    """축의금을 엑셀로 내보낸다."""
    groom_guests = Guest.query.filter_by(side='groom').order_by(Guest.id).all()
    bride_guests = Guest.query.filter_by(side='bride').order_by(Guest.id).all()

    wb = Workbook()
    ws = wb.active
    ws.title = "축의금 관리"

    title_font = Font(size=16, bold=True)
    header_font = Font(bold=True)
    align_center = Alignment(horizontal="center", vertical="center")

    ws.merge_cells("A1:H1")
    ws["A1"] = "축의금 관리"
    ws["A1"].font = title_font
    ws["A1"].alignment = align_center

    ws.merge_cells("A3:D3")
    ws.merge_cells("E3:H3")
    ws["A3"] = "신랑측"
    ws["E3"] = "신부측"
    ws["A3"].font = header_font
    ws["E3"].font = header_font
    ws["A3"].alignment = align_center
    ws["E3"].alignment = align_center

    headers = ["번호", "이름", "금액", "메모"]
    header_row = 4
    for idx, header in enumerate(headers, start=1):
        cell_groom = ws.cell(row=header_row, column=idx, value=header)
        cell_bride = ws.cell(row=header_row, column=idx + 4, value=header)
        for cell in (cell_groom, cell_bride):
            cell.font = header_font
            cell.alignment = align_center

    start_row = 5
    max_rows = max(len(groom_guests), len(bride_guests))

    for i in range(max_rows):
        row_idx = start_row + i

        if i < len(groom_guests):
            groom = groom_guests[i]
            ws.cell(row=row_idx, column=1, value=i + 1).alignment = align_center
            ws.cell(row=row_idx, column=2, value=groom.name).alignment = align_center
            amount_cell = ws.cell(row=row_idx, column=3, value=groom.amount)
            amount_cell.number_format = "#,##0"
            amount_cell.alignment = align_center
            ws.cell(row=row_idx, column=4, value=groom.note or "").alignment = align_center

        if i < len(bride_guests):
            bride = bride_guests[i]
            ws.cell(row=row_idx, column=5, value=i + 1).alignment = align_center
            ws.cell(row=row_idx, column=6, value=bride.name).alignment = align_center
            amount_cell = ws.cell(row=row_idx, column=7, value=bride.amount)
            amount_cell.number_format = "#,##0"
            amount_cell.alignment = align_center
            ws.cell(row=row_idx, column=8, value=bride.note or "").alignment = align_center

    widths = {
        "A": 8, "B": 15, "C": 12, "D": 25,
        "E": 8, "F": 15, "G": 12, "H": 25,
    }
    for col, width in widths.items():
        ws.column_dimensions[col].width = width

    ws.freeze_panes = "A5"

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    return send_file(
        output,
        as_attachment=True,
        download_name="guests.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


# PDF에 사용할 한글 폰트를 찾아본다.
def _font_path() -> Optional[str]:
    candidates = [
        Path("C:/Windows/Fonts/malgun.ttf"),
        Path("C:/Windows/Fonts/malgunbd.ttf"),
        Path("/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf"),
        Path("/Library/Fonts/Arial Unicode.ttf"),
    ]
    for path in candidates:
        if path.exists():
            return str(path)
    return None


@guests_bp.route('/export/pdf')
def export_pdf():
    """총합/차액과 전체·측별 Top 5를 PDF로 요약한다."""
    groom_total, bride_total = totals()
    total = groom_total + bride_total
    diff = groom_total - bride_total
    top_all = Guest.query.order_by(Guest.amount.desc()).limit(5).all()
    top_groom = Guest.query.filter_by(side='groom').order_by(Guest.amount.desc()).limit(5).all()
    top_bride = Guest.query.filter_by(side='bride').order_by(Guest.amount.desc()).limit(5).all()

    font_path = _font_path()
    pdf = FPDF()
    pdf.add_page()

    if font_path:
        pdf.add_font("Body", fname=font_path, uni=True)
        pdf.set_font("Body", size=14)
    else:
        pdf.set_font("Helvetica", size=14)

    pdf.cell(0, 10, txt="축의금 요약 리포트", ln=True, align="C")
    pdf.set_font_size(12)
    pdf.ln(4)
    pdf.cell(0, 8, txt=f"신랑측 합계: {groom_total:,} 원", ln=True)
    pdf.cell(0, 8, txt=f"신부측 합계: {bride_total:,} 원", ln=True)
    pdf.cell(0, 8, txt=f"총 합계: {total:,} 원", ln=True)
    pdf.cell(0, 8, txt=f"차액(신랑-신부): {diff:,} 원", ln=True)

    def draw_table(title: str, rows):
        pdf.ln(6)
        pdf.set_font_size(13)
        pdf.cell(0, 8, txt=title, ln=True)
        pdf.set_font_size(11)
        pdf.ln(2)
        col_widths = [15, 45, 30, 25, 70]  # 순위, 이름, 금액, 소속, 메모
        headers = ["#", "이름", "금액", "소속", "메모"]
        pdf.set_fill_color(240, 240, 240)
        for w, h in zip(col_widths, headers):
            pdf.cell(w, 8, txt=h, border=1, fill=True, align="C")
        pdf.ln(8)
        for idx, guest in enumerate(rows, start=1):
            side_text = "신랑측" if guest.side == "groom" else "신부측"
            pdf.cell(col_widths[0], 8, txt=str(idx), border=1, align="C")
            pdf.cell(col_widths[1], 8, txt=str(guest.name), border=1)
            pdf.cell(col_widths[2], 8, txt=f"{guest.amount:,}", border=1, align="R")
            pdf.cell(col_widths[3], 8, txt=side_text, border=1, align="C")
            pdf.cell(col_widths[4], 8, txt=str(guest.note or ""), border=1)
            pdf.ln(8)

    draw_table("전체 Top 5", top_all)
    draw_table("신랑측 Top 5", top_groom)
    draw_table("신부측 Top 5", top_bride)

    pdf_bytes = pdf.output(dest="S").encode("latin-1")
    output = io.BytesIO(pdf_bytes)

    return send_file(
        output,
        as_attachment=True,
        download_name="guests_summary.pdf",
        mimetype="application/pdf",
    )
